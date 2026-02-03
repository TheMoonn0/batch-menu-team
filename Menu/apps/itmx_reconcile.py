# -*- coding: utf-8 -*-
"""
ITMX Reconcile Tool
- อ่านไฟล์ ACQ และ ISS จากไฟล์ ZIP (เฉพาะ Folder ที่มี Faat หรือ ATMI)
- ตัดข้อมูลตามตำแหน่งที่กำหนด
- เปรียบเทียบข้อมูลระหว่าง 2 Source โดยใช้สูตร Excel
- Sheet ACQ (ถอน) / ISS (ถอน) สำหรับแถวที่ขึ้นต้นด้วยเลข 1
- Sheet ACQ (โอน) / ISS (โอน) สำหรับแถวที่ไม่ใช่เลข 1
- Sort ข้อมูลก่อนนำมาเปรียบเทียบ
- Conditional Formatting สีเขียว/แดง
"""

import streamlit as st
import os
import re
import io
import zipfile
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Regex pattern สำหรับลบ illegal characters ที่ Excel ไม่รองรับ
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

# กำหนดตำแหน่งการตัดข้อมูล (1-indexed) - สำหรับ ACQ/ISS (ถอน) Record Type 1
COLUMN_POSITIONS = [
    {"name": "Detail Record Type", "from": 1, "to": 1, "len": 1},
    {"name": "Issuer Bank Code", "from": 2, "to": 3, "len": 2},
    {"name": "PAN", "from": 4, "to": 22, "len": 19},
    {"name": "Acquirer Bank Code", "from": 23, "to": 24, "len": 2},
    {"name": "ATM Terminal ID", "from": 25, "to": 40, "len": 16},
    {"name": "Receipt Sequence Number", "from": 41, "to": 46, "len": 6},
    {"name": "Transaction Date", "from": 47, "to": 52, "len": 6},
    {"name": "Transaction Time", "from": 53, "to": 58, "len": 6},
    {"name": "Transaction Type", "from": 59, "to": 64, "len": 6},
    {"name": "Reverse Flag", "from": 65, "to": 66, "len": 2},
    {"name": "Form-account Number", "from": 67, "to": 85, "len": 19},
    {"name": "To-account Number", "from": 86, "to": 104, "len": 19},
    {"name": "Amount1", "from": 105, "to": 114, "len": 10},
    {"name": "Amount2", "from": 115, "to": 124, "len": 10},
    {"name": "Terminal State", "from": 125, "to": 128, "len": 4},
    {"name": "Response Code", "from": 129, "to": 131, "len": 3},
    {"name": "Fee Charge", "from": 132, "to": 137, "len": 6},
    {"name": "Flag", "from": 138, "to": 150, "len": 13},
]

# กำหนดตำแหน่งการตัดข้อมูล (1-indexed) - สำหรับ ACQ (โอน) Record Type 2,3,...
COLUMN_POSITIONS_TRANSFER = [
    {"name": "Detail Record Type", "from": 1, "to": 1, "len": 1},
    {"name": "From Bank Code", "from": 2, "to": 3, "len": 2},
    {"name": "PAN", "from": 4, "to": 22, "len": 19},
    {"name": "Acquirer Bank Code", "from": 23, "to": 24, "len": 2},
    {"name": "Terminal ID", "from": 25, "to": 40, "len": 16},
    {"name": "Receipt Sequence Number", "from": 41, "to": 46, "len": 6},
    {"name": "Transaction Date", "from": 47, "to": 52, "len": 6},
    {"name": "Transaction Time", "from": 53, "to": 58, "len": 6},
    {"name": "Switching Date", "from": 59, "to": 62, "len": 4},
    {"name": "Transaction Type", "from": 63, "to": 68, "len": 6},
    {"name": "Reverse Flag", "from": 69, "to": 70, "len": 2},
    {"name": "Form-account Number", "from": 71, "to": 89, "len": 19},
    {"name": "To-account Number", "from": 90, "to": 108, "len": 19},
    {"name": "Amount1", "from": 109, "to": 120, "len": 12},
    {"name": "Terminal State", "from": 121, "to": 124, "len": 4},
    {"name": "Response Code", "from": 125, "to": 127, "len": 3},
    {"name": "Issuer Fee", "from": 128, "to": 133, "len": 6},
    {"name": "To Account Fee", "from": 134, "to": 138, "len": 5},
    {"name": "Acquirer Fee", "from": 139, "to": 143, "len": 5},
    {"name": "To Bank Code", "from": 144, "to": 145, "len": 2},
    {"name": "DR/CR Flag", "from": 146, "to": 146, "len": 1},
    {"name": "Reserve", "from": 147, "to": 150, "len": 4},
]


def sanitize_for_excel(value):
    """ลบ illegal characters ที่ Excel ไม่รองรับ"""
    if value is None:
        return ""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def find_files_by_keyword(directory, keyword):
    """ค้นหาไฟล์ที่มี keyword ในชื่อ (ไม่รวมไฟล์ Control Report)"""
    files = []
    if os.path.exists(directory):
        for filename in os.listdir(directory):
            # ข้ามไฟล์ Control Report (F133.C*)
            if "F133.C" in filename.upper():
                continue
            if keyword.upper() in filename.upper():
                files.append(os.path.join(directory, filename))
    return files


def read_and_parse_file(filepath, record_type='1', column_positions=None):
    """อ่านไฟล์และตัดข้อมูลตามตำแหน่งที่กำหนด (ไม่รวมแถวแรกและแถวสุดท้าย)
    record_type: '1' สำหรับถอน, 'not1' สำหรับโอน (ทุกเลขที่ไม่ใช่ 1)
    column_positions: ตำแหน่ง columns ที่จะใช้ parse
    """
    if column_positions is None:
        column_positions = COLUMN_POSITIONS
    
    data = []
    
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
    except Exception as e:
        return data
    
    # ข้ามแถวแรกและแถวสุดท้าย
    if len(lines) <= 2:
        return data
    
    data_lines = lines[1:-1]  # ตัดแถวแรกและแถวสุดท้าย
    
    for line_num, line in enumerate(data_lines, start=2):
        # ลบ newline characters
        line = line.rstrip('\r\n')
        
        # ถ้าบรรทัดว่างหรือเป็น EOF character ให้ข้าม
        if not line or line == '\x1a':
            continue
        
        # ตรวจสอบ record type
        first_char = line[0] if line else ''
        
        if record_type == '1':
            # เฉพาะแถวที่ขึ้นต้นด้วยเลข 1
            if first_char != '1':
                continue
        elif record_type == 'not1':
            # เฉพาะแถวที่ขึ้นต้นด้วยตัวเลข 2-8 (ไม่รวม 0,1,9 ซึ่งเป็น header/trailer)
            if not first_char.isdigit() or first_char in ('0', '1', '9'):
                continue
        else:
            # กรณีระบุ record_type อื่นๆ
            if first_char != record_type:
                continue
        
        row = {"LINE_NUM": line_num}
        
        for col in column_positions:
            # Python string is 0-indexed, แต่ตำแหน่งที่กำหนดเป็น 1-indexed
            start_idx = col["from"] - 1
            end_idx = col["to"]
            
            # ตัดข้อมูลตามตำแหน่ง
            if len(line) >= end_idx:
                value = line[start_idx:end_idx]
            elif len(line) > start_idx:
                value = line[start_idx:]
            else:
                value = ""
            
            row[col["name"]] = value
        
        # เก็บ raw line เพื่อใช้เปรียบเทียบและ sort
        row["RAW_LINE"] = line
        
        data.append(row)
    
    return data


def sort_data(data):
    """Sort ข้อมูลตาม Issuer Bank Code, PAN, ATM Terminal ID, And Sequence Num, Transaction Date, Timeout of Time"""
    return sorted(data, key=lambda x: (
        x.get("Issuer Bank Code", "") or x.get("From Bank Code", ""),
        x.get("PAN", ""),
        x.get("ATM Terminal ID", "") or x.get("Terminal ID", ""),
        x.get("Receipt Sequence Number", ""),
        x.get("Transaction Date", ""),
        x.get("Transaction Time", "")
    ))


def get_record_key(record, column_positions):
    """สร้าง key สำหรับ match records"""
    # ใช้ PAN, Terminal ID, Sequence Number, Date, Time เป็น key
    key_fields = []
    for col in column_positions:
        name = col['name']
        if name in ['PAN', 'ATM Terminal ID', 'Terminal ID', 'Receipt Sequence Number', 
                    'Transaction Date', 'Transaction Time']:
            key_fields.append(record.get(name, "").strip())
    return tuple(key_fields)


def align_data_by_key(faa_data, atmi_data, column_positions):
    """จับคู่ข้อมูล FAA และ ATMI ตาม key และเรียงให้ตรงกัน
    Return: (aligned_faa, aligned_atmi) - lists ที่มี None หากไม่มี match
    """
    # สร้าง dictionary ของ ATMI records โดยใช้ key
    atmi_by_key = {}
    for record in atmi_data:
        key = get_record_key(record, column_positions)
        if key not in atmi_by_key:
            atmi_by_key[key] = []
        atmi_by_key[key].append(record)
    
    aligned_faa = []
    aligned_atmi = []
    
    # จับคู่ FAA records กับ ATMI records
    for faa_record in faa_data:
        key = get_record_key(faa_record, column_positions)
        
        if key in atmi_by_key and atmi_by_key[key]:
            # พบ match - เอา ATMI record แรกที่ยังไม่ถูกใช้
            atmi_record = atmi_by_key[key].pop(0)
            aligned_faa.append(faa_record)
            aligned_atmi.append(atmi_record)
        else:
            # ไม่พบ match - ใส่ FAA record กับ None
            aligned_faa.append(faa_record)
            aligned_atmi.append(None)
    
    # เพิ่ม ATMI records ที่เหลือ (ไม่มี FAA match)
    for key, records in atmi_by_key.items():
        for record in records:
            aligned_faa.append(None)
            aligned_atmi.append(record)
    
    return aligned_faa, aligned_atmi


def create_sheet_content(ws, faatbatch_data, atmi_data, faa_filename, atmi_filename, column_positions=None):
    """สร้างเนื้อหาใน Sheet พร้อมสูตร Excel เปรียบเทียบ และ Conditional Formatting"""
    
    if column_positions is None:
        column_positions = COLUMN_POSITIONS
    
    # จัดเรียงข้อมูลให้ตรงกันตาม key
    aligned_faa, aligned_atmi = align_data_by_key(faatbatch_data, atmi_data, column_positions)
    
    # สไตล์
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    same_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    current_row = 1
    num_data_cols = len(column_positions)
    last_col_letter = get_column_letter(num_data_cols)
    total_rows = len(aligned_faa)
    
    # ========== FAA SECTION HEADER ==========
    cell = ws.cell(row=current_row, column=1, value=f"FAA - {faa_filename}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = left_align
    cell.border = thin_border
    
    # Merge cells for header to show full filename
    ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
    current_row += 1
    
    # ========== FAA COLUMN HEADERS ==========
    for col_idx, col in enumerate(column_positions, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1
    
    # ========== FAA DATA ROWS (aligned) ==========
    faa_start_row = current_row
    for row_idx, data_row in enumerate(aligned_faa):
        for col_idx, col in enumerate(column_positions, 1):
            if data_row is not None:
                value = sanitize_for_excel(data_row.get(col['name'], ""))
            else:
                value = ""  # ว่างสำหรับ record ที่ไม่ match
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
        current_row += 1
    
    faa_end_row = current_row - 1
    
    # ========== ATMI SECTION HEADER ==========
    cell = ws.cell(row=current_row, column=1, value=f"ATMI - {atmi_filename}")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = left_align
    cell.border = thin_border
    
    # Merge cells for header to show full filename
    ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
    current_row += 1
    
    # ========== ATMI COLUMN HEADERS ==========
    for col_idx, col in enumerate(column_positions, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1
    
    # ========== ATMI DATA ROWS (aligned) ==========
    atmi_start_row = current_row
    for row_idx, data_row in enumerate(aligned_atmi):
        for col_idx, col in enumerate(column_positions, 1):
            if data_row is not None:
                value = sanitize_for_excel(data_row.get(col['name'], ""))
            else:
                value = ""  # ว่างสำหรับ record ที่ไม่ match
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
        current_row += 1
    
    atmi_end_row = current_row - 1
    
    # ========== EMPTY ROW ==========
    current_row += 1
    
    # ========== COMPARISON SECTION WITH FORMULAS ==========
    compare_start_row = current_row
    
    for i in range(total_rows):
        faa_data_row = faa_start_row + i  # แถวข้อมูล FAA ที่ตรงกัน
        atmi_data_row = atmi_start_row + i  # แถวข้อมูล ATMI ที่ตรงกัน
        
        # Data Columns with comparison formulas (start from column 1)
        for col_idx, col in enumerate(column_positions, 1):
            col_letter = get_column_letter(col_idx)
            
            # สูตร: ถ้า FAA == ATMI แสดง "same" ไม่งั้นแสดง "diff"
            formula = f'=IF({col_letter}{faa_data_row}={col_letter}{atmi_data_row},"same","diff")'
            
            cell = ws.cell(row=current_row, column=col_idx, value=formula)
            cell.border = thin_border
        
        current_row += 1
    
    compare_end_row = current_row - 1
    
    # ========== CONDITIONAL FORMATTING ==========
    if total_rows > 0:
        compare_range = f"A{compare_start_row}:{last_col_letter}{compare_end_row}"
        
        # Rule สำหรับ "same" - สีเขียว
        same_rule = FormulaRule(
            formula=[f'A{compare_start_row}="same"'],
            fill=same_fill
        )
        ws.conditional_formatting.add(compare_range, same_rule)
        
        # Rule สำหรับ "diff" - สีแดง
        diff_rule = FormulaRule(
            formula=[f'A{compare_start_row}="diff"'],
            fill=diff_fill
        )
        ws.conditional_formatting.add(compare_range, diff_rule)
    
    # ========== ปรับความกว้าง Column ==========
    for i, col in enumerate(column_positions, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(len(col['name']) + 2, col['len'] + 2, 8)


def process_file_type(faatbatch_files, atmi_files, record_type, column_positions=None):
    """ประมวลผลไฟล์และคืนค่า data พร้อม filename"""
    if column_positions is None:
        column_positions = COLUMN_POSITIONS
    
    faa_filename = ""
    atmi_filename = ""
    
    # อ่านและ Parse ข้อมูล FAA
    all_faatbatch_data = []
    for filepath in faatbatch_files:
        data = read_and_parse_file(filepath, record_type, column_positions)
        if data:
            faa_filename = os.path.basename(filepath)
        all_faatbatch_data.extend(data)
    
    # อ่านและ Parse ข้อมูล ATMI
    all_atmi_data = []
    for filepath in atmi_files:
        data = read_and_parse_file(filepath, record_type, column_positions)
        if data:
            atmi_filename = os.path.basename(filepath)
        all_atmi_data.extend(data)
    
    # Sort ข้อมูล
    all_faatbatch_data = sort_data(all_faatbatch_data)
    all_atmi_data = sort_data(all_atmi_data)
    
    return all_faatbatch_data, all_atmi_data, faa_filename, atmi_filename


def process_itmx_data(faat_dir, atmi_dir):
    """หลักสำหรับประมวลผล ITMX data และคืนค่า Excel bytes"""
    
    # ค้นหาไฟล์ ACQ
    faatbatch_acq_files = find_files_by_keyword(faat_dir, "ACQ")
    atmi_acq_files = find_files_by_keyword(atmi_dir, "ACQ")
    
    # ค้นหาไฟล์ ISS
    faatbatch_iss_files = find_files_by_keyword(faat_dir, "ISS")
    atmi_iss_files = find_files_by_keyword(atmi_dir, "ISS")
    
    # ประมวลผล ACQ (ถอน) - Record Type 1
    acq_faa_1, acq_atmi_1, acq_faa_fn_1, acq_atmi_fn_1 = process_file_type(
        faatbatch_acq_files, atmi_acq_files, '1'
    )
    
    # ประมวลผล ACQ (โอน) - ทุก Record Type ที่ไม่ใช่ 1 (ใช้ COLUMN_POSITIONS_TRANSFER)
    acq_faa_not1, acq_atmi_not1, acq_faa_fn_not1, acq_atmi_fn_not1 = process_file_type(
        faatbatch_acq_files, atmi_acq_files, 'not1', COLUMN_POSITIONS_TRANSFER
    )
    
    # ประมวลผล ISS (ถอน) - Record Type 1
    iss_faa_1, iss_atmi_1, iss_faa_fn_1, iss_atmi_fn_1 = process_file_type(
        faatbatch_iss_files, atmi_iss_files, '1'
    )
    
    # ประมวลผล ISS (โอน) - ทุก Record Type ที่ไม่ใช่ 1
    iss_faa_not1, iss_atmi_not1, iss_faa_fn_not1, iss_atmi_fn_not1 = process_file_type(
        faatbatch_iss_files, atmi_iss_files, 'not1'
    )
    
    # สร้าง Excel Report
    wb = Workbook()
    first_sheet_created = False
    
    # ========== Sheet 1: ACQ (ถอน) ==========
    if acq_faa_1 or acq_atmi_1:
        if not first_sheet_created:
            ws = wb.active
            ws.title = "ACQ (ถอน)"
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title="ACQ (ถอน)")
        create_sheet_content(ws, acq_faa_1, acq_atmi_1, acq_faa_fn_1, acq_atmi_fn_1)
    
    # ========== Sheet 2: ACQ (โอน) - ใช้ COLUMN_POSITIONS_TRANSFER ==========
    if acq_faa_not1 or acq_atmi_not1:
        if not first_sheet_created:
            ws = wb.active
            ws.title = "ACQ (โอน)"
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title="ACQ (โอน)")
        create_sheet_content(ws, acq_faa_not1, acq_atmi_not1, acq_faa_fn_not1, acq_atmi_fn_not1, COLUMN_POSITIONS_TRANSFER)
    
    # ========== Sheet 3: ISS (ถอน) ==========
    if iss_faa_1 or iss_atmi_1:
        if not first_sheet_created:
            ws = wb.active
            ws.title = "ISS (ถอน)"
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title="ISS (ถอน)")
        create_sheet_content(ws, iss_faa_1, iss_atmi_1, iss_faa_fn_1, iss_atmi_fn_1)
    
    # ========== Sheet 4: ISS (โอน) ==========
    if iss_faa_not1 or iss_atmi_not1:
        if not first_sheet_created:
            ws = wb.active
            ws.title = "ISS (โอน)"
            first_sheet_created = True
        else:
            ws = wb.create_sheet(title="ISS (โอน)")
        create_sheet_content(ws, iss_faa_not1, iss_atmi_not1, iss_faa_fn_not1, iss_atmi_fn_not1)
    
    # ลบ default Sheet ถ้าไม่ได้ใช้
    if not first_sheet_created:
        return None, "ไม่พบข้อมูล ACQ/ISS ใน Folder ที่กำหนด"
    
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, None


def find_faat_atmi_folders(temp_dir):
    """ค้นหา Folder ที่มี 'Faat' หรือ 'ATMI' ในชื่อ"""
    faat_dir = None
    atmi_dir = None
    
    for root, dirs, files in os.walk(temp_dir):
        for d in dirs:
            dir_lower = d.lower()
            if 'faat' in dir_lower and faat_dir is None:
                faat_dir = os.path.join(root, d)
            if 'atmi' in dir_lower and atmi_dir is None:
                atmi_dir = os.path.join(root, d)
    
    return faat_dir, atmi_dir


def render():
    """Render ITMX Reconcile page - เปรียบเทียบข้อมูล ITMX จาก Faat และ ATMI folders"""
    
    st.write("อัปโหลดไฟล์ ZIP ที่ประกอบด้วย Folder ที่มีชื่อว่า **Faat** และ **ATMI**")
    st.caption("ระบบจะค้นหาเฉพาะ Folder ที่มีคำว่า 'Faat' หรือ 'ATMI' ในชื่อ และค้นหาไฟล์ ACQ/ISS ภายใน")
    
    uploaded_zip = st.file_uploader("Choose a ZIP file", type="zip", key="itmx_zip")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"ITMX_Comparison_{timestamp}.xlsx"
    output_name = st.text_input("Output Excel filename", value=default_filename)
    
    if uploaded_zip:
        if st.button("🚀 Process Files", type="primary", key="itmx_process"):
            with st.spinner("Extracting & Processing..."):
                with tempfile.TemporaryDirectory() as temp_dir:
                    try:
                        # Extract ZIP
                        with zipfile.ZipFile(uploaded_zip, "r") as zip_ref:
                            zip_ref.extractall(temp_dir)
                        
                        # Find Faat and ATMI folders
                        faat_dir, atmi_dir = find_faat_atmi_folders(temp_dir)
                        
                        if not faat_dir:
                            st.error("❌ ไม่พบ Folder ที่มีคำว่า 'Faat' ใน ZIP")
                            return
                        
                        if not atmi_dir:
                            st.error("❌ ไม่พบ Folder ที่มีคำว่า 'ATMI' ใน ZIP")
                            return
                        
                        st.info(f"📍 Found Faat Folder: {os.path.basename(faat_dir)}")
                        st.info(f"📍 Found ATMI Folder: {os.path.basename(atmi_dir)}")
                        
                        # Process data
                        excel_file, error_msg = process_itmx_data(faat_dir, atmi_dir)
                        
                        if error_msg:
                            st.error(f"❌ {error_msg}")
                        else:
                            st.success("✅ Processing Complete!")
                            st.download_button(
                                label="📥 Download ITMX Comparison Excel",
                                data=excel_file,
                                file_name=output_name.strip() or default_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                    
                    except zipfile.BadZipFile:
                        st.error("❌ ZIP ไฟล์เสียหรือเปิดไม่ได้")
                    except Exception as e:
                        st.error(f"❌ Error during processing: {e}")
