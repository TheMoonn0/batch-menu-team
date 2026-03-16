import io
import os
import re
import sys
import traceback
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Config ---
DEFAULT_SOURCE_FOLDER_NAME = "CDM"

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_MENU_DIR = os.path.dirname(_SCRIPT_DIR)
DEFAULT_TLF_FOLDER = os.path.join(_MENU_DIR, "Data GL")

# Keep legacy globals defined so older helper functions remain harmless if called.
source_folder = DEFAULT_SOURCE_FOLDER_NAME
tlf_filenames = []
source_folder_name = os.path.basename(os.path.normpath(source_folder))
output_filename = f"GL_{source_folder_name}.xlsx"

tlf_reserved_rows = 2
gl_reserved_rows = 10
gap_rows = 2

TLF_LABEL = "TLF or Database(ATMI)"
CSV_ENCODINGS = ["utf-8", "cp874"]
TERM_TYPE_FILTERS = {"ATM": "N", "CDM": "R"}
TLF_TIME_HEADER = "auth_tran_tim_hms"
HELPER_COLUMN_NAMES = {"_SearchKey", "_SearchKeyTime"}
SEARCH_FIXED_WIDTHS = {TLF_TIME_HEADER: 16, "Details": 12}
SEARCH_HEADER_LABELS = {"DR": "   DR", "CR": "   CR", "Details": "          Details"}
SEARCH_WIDTH_LIMITS = (10, 24)

gl_columns_letters = ["J", "K", "L", "M", "N", "P", "AM", "AN", "AZ"]
gl_base_headers = ["RC", "OC", "CH", "Product Code", "Account Code", "Tax", "DR", "CR", "Seq", "Details"]
gl_source_headers = [
    "RC",
    "OC",
    "CH",
    "Product Code",
    "Account Code",
    "Tax",
    "DR",
    "CR",
    "AZ_RAW",
]

tlf_columns_letters = [
    "F",
    "I",
    "G",
    "J",
    "K",
    "L",
    "M",
    "O",
    "V",
    "W",
    "AF",
    "AS",
    "AT",
    "AU",
    "AV",
    "AX",
    "AZ",
    "CU",
    "BH",
    "BW",
    "BX",
    "DP",
    "DQ",
    "C",
    "DU",
]


def configure_stdout() -> None:
    """Keep Thai log output readable on Windows terminals."""
    stdout = getattr(sys, "stdout", None)
    if stdout and hasattr(stdout, "reconfigure"):
        try:
            stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass


def excel_col_to_index(col_str):
    num = 0
    for char in col_str:
        if char in "0123456789":
            continue
        num = num * 26 + (ord(char.upper()) - ord("A")) + 1
    return num - 1


def convert_implied_decimal(val):
    """Convert implied decimal only when the source value starts with 00."""
    try:
        if val is None:
            return val
        val_str = str(val).strip()
        if not val_str.isdigit() or not val_str.startswith("00"):
            return val_str
        return float(val_str) / 100.0
    except Exception:
        return val


def extract_seq_num(val):
    text = str(val)
    match = re.search(r"seq_num:(\d+)", text)
    if match:
        return match.group(1)
    return str(val).strip()


def strip_d_suffix_for_tlf_sheet(name_no_ext: str):
    return re.sub(r"[-_]?D{6}.*$", "", name_no_ext, flags=re.IGNORECASE).strip()


def make_unique_sheet_name(book, desired_name: str):
    base = (desired_name or "Sheet")[:31]
    name = base
    suffix_num = 2
    while name in book.sheetnames:
        suffix = f"_{suffix_num}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        suffix_num += 1
    return name


def max_k_from_searchkey(series: pd.Series) -> int:
    max_k = 1
    try:
        k_series = series.astype(str).str.extract(r"\|(\d+)\s*$")[0]
        k_series = pd.to_numeric(k_series, errors="coerce")
        max_k_val = k_series.max()
        if pd.notna(max_k_val):
            max_k = int(max_k_val)
    except Exception:
        max_k = 1
    return max_k


def get_term_type_from_source_name(source_name: str):
    normalized_source_name = (source_name or "").strip().upper()
    return normalized_source_name, TERM_TYPE_FILTERS.get(normalized_source_name)


def to_yymmdd(value):
    if value is None:
        return None

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%y%m%d")

    stripped = str(value).strip()
    if not stripped or stripped.lower() in ["nan", "none", "null"]:
        return None

    digits = re.sub(r"[^0-9]", "", stripped)

    if len(digits) == 6:
        return digits

    if len(digits) == 8:
        first4 = int(digits[:4])
        last4 = int(digits[-4:])

        if 1900 <= first4 <= 2099:
            try:
                dt = datetime.strptime(digits, "%Y%m%d")
                return dt.strftime("%y%m%d")
            except Exception:
                pass

        if 1900 <= last4 <= 2099:
            try:
                dt = datetime.strptime(digits, "%d%m%Y")
                return dt.strftime("%y%m%d")
            except Exception:
                pass

    try:
        dt = pd.to_datetime(stripped, errors="coerce", dayfirst=True)
        if pd.isna(dt):
            return None
        return dt.strftime("%y%m%d")
    except Exception:
        return None


def pick_date_from_csv_col_c(file_path, nrows=20):
    """Pick the most frequent date from Column C and return it as YYMMDD."""
    df_col_c = None
    for encoding in CSV_ENCODINGS:
        try:
            df_col_c = pd.read_csv(
                file_path,
                header=None,
                usecols=[2],
                nrows=nrows,
                encoding=encoding,
                dtype=str,
                engine="python",
            )
            break
        except Exception:
            df_col_c = None

    if df_col_c is None or df_col_c.empty:
        return None

    date_values = [to_yymmdd(value) for value in df_col_c.iloc[:, 0].tolist()]
    date_values = [value for value in date_values if value]
    if not date_values:
        return None

    value_counts = pd.Series(date_values).value_counts()
    max_count = value_counts.max()
    candidates = [date_value for date_value, count in value_counts.items() if count == max_count]
    return max(candidates, key=lambda item: int(item))


def list_csv_files(folder_path):
    files = []
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path) and filename.lower().endswith(".csv"):
            files.append(file_path)
    files.sort(key=lambda path: os.path.basename(path).lower())
    return files


def get_col_pos_in_tlf(target_letter):
    try:
        return tlf_columns_letters.index(target_letter)
    except ValueError:
        return -1


def sanitize_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)


def make_visible_text_formula(text: str, *required_refs: str) -> str:
    escaped_text = str(text).replace('"', '""')
    empty_checks = ",".join(f'{cell_ref}=""' for cell_ref in required_refs)
    return f'=IF(OR({empty_checks}), "", "{escaped_text}")'


def read_gl_csv(file_path, gl_indices):
    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(
                file_path,
                header=None,
                usecols=gl_indices,
                encoding=encoding,
                dtype=str,
                engine="python",
            )
        except Exception:
            continue
    raise ValueError(f"ไม่สามารถอ่านไฟล์ด้วย encoding ที่รองรับได้: {file_path}")


def normalize_string_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    for column in df.columns:
        stripped = df[column].astype(str).str.strip()
        df[column] = stripped.apply(lambda value: "" if value.lower() in ("nan", "none") else value)
    return df


def build_tlf_lookup_candidates(filename: str, chosen_date):
    clean_name = re.sub(r"GL", "", filename, flags=re.IGNORECASE)
    clean_name = clean_name.replace(".csv", "").replace(".CSV", "").strip()
    fallback_lookup_name = strip_d_suffix_for_tlf_sheet(clean_name)

    candidates = []
    if chosen_date:
        candidates.append(chosen_date)
        candidates.append("D" + chosen_date)
    candidates.append(fallback_lookup_name)
    return candidates


def find_tlf_sheet(tlf_books, lookup_candidates):
    for book_filename, book in tlf_books:
        for candidate in lookup_candidates:
            if candidate and candidate in book.sheet_names:
                return candidate, book, book_filename
    return None, None, None


def prepare_tlf_dataframe(tlf_book, tlf_sheet_name, tlf_filter_val, tlf_reorder, pos_term_typ, pos_az, pos_cu, pos_af):
    tlf_df = pd.read_excel(
        tlf_book,
        sheet_name=tlf_sheet_name,
        usecols=tlf_indices,
        dtype=str,
    )
    tlf_df = tlf_df.iloc[:, tlf_reorder]
    tlf_df = normalize_string_dataframe(tlf_df)

    if tlf_filter_val and pos_term_typ != -1 and pos_term_typ < len(tlf_df.columns):
        term_type_col = tlf_df.iloc[:, pos_term_typ].astype(str).str.strip()
        tlf_df = tlf_df[term_type_col == tlf_filter_val].reset_index(drop=True)

    if pos_az != -1 and pos_az < len(tlf_df.columns):
        tlf_df.iloc[:, pos_az] = tlf_df.iloc[:, pos_az].apply(convert_implied_decimal)
    if pos_cu != -1 and pos_cu < len(tlf_df.columns):
        tlf_df.iloc[:, pos_cu] = tlf_df.iloc[:, pos_cu].apply(convert_implied_decimal)

    max_k_tlf = 1
    effective_tlf_reserved_rows = tlf_reserved_rows
    if not tlf_df.empty and pos_af != -1 and pos_af < len(tlf_df.columns):
        # Build stable lookup keys for Search formulas.
        search_col = tlf_df.iloc[:, pos_af].astype(str).str.strip()
        seq_occurrence = tlf_df.groupby(search_col).cumcount() + 1
        tlf_df["_SearchKey"] = "T|" + search_col + "|" + seq_occurrence.astype(str)

        if TLF_TIME_HEADER in tlf_df.columns:
            time_col = tlf_df[TLF_TIME_HEADER].astype(str).str.strip()
        else:
            time_col = pd.Series([""] * len(tlf_df), index=tlf_df.index)

        time_occurrence = pd.DataFrame(
            {"seq": search_col, "time": time_col},
            index=tlf_df.index,
        ).groupby(["seq", "time"]).cumcount() + 1
        tlf_df["_SearchKeyTime"] = "TT|" + search_col + "|" + time_col + "|" + time_occurrence.astype(str)
        max_k_tlf = max_k_from_searchkey(tlf_df["_SearchKey"])
        effective_tlf_reserved_rows = max(tlf_reserved_rows, max_k_tlf)
    else:
        tlf_df["_SearchKey"] = ""
        tlf_df["_SearchKeyTime"] = ""

    return tlf_df, effective_tlf_reserved_rows, max_k_tlf


def prepare_gl_dataframe(file_path):
    gl_df = read_gl_csv(file_path, gl_indices)

    if len(gl_df.columns) == len(gl_source_headers):
        gl_df.columns = gl_source_headers

    gl_df["Details"] = gl_df["AZ_RAW"]
    gl_df["Seq"] = gl_df["AZ_RAW"].apply(extract_seq_num).astype(str).str.strip()

    gl_df["RC"] = gl_df["RC"].astype(str).str.strip()
    gl_df["CH"] = gl_df["CH"].astype(str).str.strip()
    gl_df["DR"] = pd.to_numeric(gl_df["DR"], errors="coerce").fillna(0)
    gl_df["CR"] = pd.to_numeric(gl_df["CR"], errors="coerce").fillna(0)

    gl_df = gl_df[gl_base_headers]
    gl_df = gl_df.sort_values(
        by=["CH", "RC", "OC", "Product Code"],
        ascending=[True, True, True, True],
    )

    if not gl_df.empty:
        search_col_gl = gl_df["Seq"].astype(str)
        gl_df["_SearchKey"] = "G|" + search_col_gl + "|" + (gl_df.groupby(search_col_gl).cumcount() + 1).astype(str)

    max_k_gl = 1
    if not gl_df.empty and "_SearchKey" in gl_df.columns:
        max_k_gl = max_k_from_searchkey(gl_df["_SearchKey"])

    effective_gl_reserved_rows = max(gl_reserved_rows, max_k_gl)
    return gl_df, effective_gl_reserved_rows, max_k_gl


def write_tlf_section(writer, sheet_name, worksheet, tlf_df, start_row):
    worksheet.cell(row=start_row, column=1, value=TLF_LABEL).font = Font(bold=True, italic=True)
    tlf_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=start_row,
        index=False,
    )

    tlf_display_cols = [column for column in tlf_df.columns if column not in HELPER_COLUMN_NAMES]
    tlf_data_end = start_row + 1 + len(tlf_df)
    # Search reads helper keys from the raw sheet instead of recalculating them.
    tlf_key_col_letter = get_column_letter(tlf_df.columns.get_loc("_SearchKey") + 1)
    tlf_time_key_col_letter = get_column_letter(tlf_df.columns.get_loc("_SearchKeyTime") + 1)

    for row in range(start_row + 1, tlf_data_end + 1):
        for col in range(1, len(tlf_df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if row == start_row + 1:
                cell.alignment = align_center
                cell.font = header_font
            else:
                cell.alignment = align_right if isinstance(cell.value, (int, float)) else align_center
                if col == 9:
                    cell.number_format = "@"

    next_row = start_row + len(tlf_df) + 3
    return tlf_display_cols, tlf_key_col_letter, tlf_time_key_col_letter, next_row


def write_gl_section(writer, sheet_name, worksheet, gl_df, start_row):
    worksheet.cell(row=start_row, column=1, value="--- Raw ATMI Data ---").font = Font(bold=True, italic=True)
    gl_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=start_row,
        index=False,
    )

    gl_display_cols = [column for column in gl_df.columns if column not in HELPER_COLUMN_NAMES]
    gl_data_end = start_row + 1 + len(gl_df)
    gl_key_col_letter = get_column_letter(gl_df.columns.get_loc("_SearchKey") + 1)

    for row in range(start_row + 1, gl_data_end + 1):
        for col in range(1, len(gl_df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            if row == start_row + 1:
                cell.alignment = align_center
                cell.font = header_font
            else:
                if col in [7, 8]:
                    cell.alignment = align_right
                    cell.number_format = "#,##0.00"
                elif col == gl_base_headers.index("Details") + 1:
                    cell.alignment = align_left
                    cell.number_format = "@"
                else:
                    cell.alignment = align_center

                if col in [
                    gl_base_headers.index("Seq") + 1,
                    gl_base_headers.index("Details") + 1,
                ]:
                    cell.number_format = "@"

    return gl_display_cols, gl_key_col_letter


def update_max_width(col_widths, df, start_col_idx=1, skip_cols=None):
    skip_cols = set(skip_cols or [])
    for offset, col_name in enumerate(df.columns):
        if col_name in skip_cols:
            continue

        current_idx = start_col_idx + offset
        max_len = len(str(col_name))
        if not df.empty:
            try:
                data_len = df[col_name].astype(str).map(len).max()
                if pd.notna(data_len):
                    max_len = max(max_len, data_len)
            except Exception:
                pass

        existing_width = col_widths.get(current_idx, 0)
        col_widths[current_idx] = max(existing_width, max_len + 3)


def apply_column_widths(worksheet, tlf_df, tlf_display_cols, gl_df):
    col_widths = {}

    if not tlf_df.empty:
        update_max_width(col_widths, tlf_df, start_col_idx=1)
        for idx, col_name in enumerate(tlf_display_cols, 1):
            header_len = len(str(col_name)) + 3
            existing_width = col_widths.get(idx, 0)
            col_widths[idx] = max(existing_width, header_len)

    if not gl_df.empty:
        update_max_width(col_widths, gl_df, start_col_idx=1, skip_cols={"Details"})

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        final_width = max(12, min(width, 60))
        worksheet.column_dimensions[col_letter].width = final_width

    if "Details" in gl_df.columns:
        details_col_idx = gl_df.columns.get_loc("Details") + 1
        details_col_letter = get_column_letter(details_col_idx)
        worksheet.column_dimensions[details_col_letter].width = 12


def calculate_search_display_widths(df, display_cols):
    widths = {}
    min_width, max_width = SEARCH_WIDTH_LIMITS

    for idx, col_name in enumerate(display_cols, 1):
        if col_name in SEARCH_FIXED_WIDTHS:
            continue

        max_len = len(str(col_name).strip())
        if col_name in df.columns and not df.empty:
            try:
                series = df[col_name].astype(str).replace({"nan": "", "None": "", "none": ""})
                data_len = series.map(len).max()
                if pd.notna(data_len):
                    max_len = max(max_len, int(data_len))
            except Exception:
                pass

        widths[idx] = max(min_width, min(max_len + 2, max_width))

    return widths


def merge_position_widths(*width_maps):
    merged = {}
    for width_map in width_maps:
        for idx, width in width_map.items():
            merged[idx] = max(merged.get(idx, 0), width)
    return merged


def add_conditional_style(ws, cell_range, formula, *, font=None, fill=None, border=None, alignment=None):
    dxf = DifferentialStyle(font=font, fill=fill, border=border, alignment=alignment)
    ws.conditional_formatting.add(
        cell_range,
        Rule(type="expression", dxf=dxf, stopIfTrue=False, formula=[formula]),
    )


def first_section_info(date_sheets_info, display_key):
    for info in date_sheets_info:
        if info.get(display_key):
            return info
    return date_sheets_info[0]


def create_search_sheet(writer, date_sheets_info):
    if not date_sheets_info:
        return

    ws = writer.book.create_sheet("Search")
    writer.sheets["Search"] = ws

    tlf_info = first_section_info(date_sheets_info, "tlf_display_cols")
    gl_info = first_section_info(date_sheets_info, "gl_display_cols")
    tlf_display_cols = tlf_info.get("tlf_display_cols", [])
    gl_display_cols = gl_info.get("gl_display_cols", [])
    max_tlf_rows = max((info["effective_tlf_reserved_rows"] for info in date_sheets_info), default=0)
    max_gl_rows = max((info["effective_gl_reserved_rows"] for info in date_sheets_info), default=0)

    search_ui_start_row = 1

    ws[f"A{search_ui_start_row}"] = "🔍 ค้นหาข้อมูล SEQ"
    ws[f"A{search_ui_start_row}"].font = Font(bold=True, size=14)
    ws[f"A{search_ui_start_row}"].alignment = Alignment(horizontal="left")

    ws.cell(row=3, column=1, value="📅 Date:").font = Font(bold=True, size=12)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="right")

    date_cell = ws.cell(row=3, column=2)
    date_cell.fill = search_fill
    date_cell.border = thin_border
    date_cell.alignment = align_center
    date_cell.number_format = "@"

    available_dates = [info["sheet_name"] for info in date_sheets_info]
    dv = DataValidation(type="list", formula1=f'"{",".join(available_dates)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(date_cell)
    if len(available_dates) == 1:
        date_cell.value = available_dates[0]

    ws.cell(row=4, column=1, value="🔢 SEQ:").font = Font(bold=True, size=12)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="right")
    seq_cell = ws.cell(row=4, column=2)
    seq_cell.fill = search_fill
    seq_cell.border = thin_border
    seq_cell.alignment = align_center
    seq_cell.number_format = "@"

    ws.cell(row=5, column=1, value="🕒 auth_tran_tim_hms:").font = Font(bold=True, size=12)
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right")
    time_cell = ws.cell(row=5, column=2)
    time_cell.fill = search_fill
    time_cell.border = thin_border
    time_cell.alignment = align_center
    time_cell.number_format = "@"

    input_cell_ref = "$B$4"
    date_ref = "$B$3"
    time_ref = "$B$5"
    # Hidden helper cells keep Search dynamic without using macros.
    helper_visible_count_cell = "AX1"
    helper_visible_count_ref = "$AX$1"
    helper_atmi_title_row_cell = "AX2"
    helper_atmi_title_row_ref = "$AX$2"
    helper_atmi_header_row_cell = "AX3"
    helper_atmi_header_row_ref = "$AX$3"
    helper_match_col_letter = "AY"
    helper_time_col_letter = "AZ"
    helper_gl_visible_count_cell = "BA1"
    helper_gl_visible_count_ref = "$BA$1"
    tlf_last_col_letter = get_column_letter(max(len(tlf_display_cols), 1))
    gl_last_col_letter = get_column_letter(max(len(gl_display_cols), 1))
    last_result_col_idx = max(len(tlf_display_cols), len(gl_display_cols), 1)

    report_row = 7
    tlf_header_row = report_row + 1
    tlf_data_start_row = report_row + 2

    if tlf_display_cols:
        ws[f"A{report_row}"] = make_visible_text_formula(TLF_LABEL, date_ref, input_cell_ref)
        ws[f"A{report_row}"].font = title_font

        for current_col_idx, col_name in enumerate(tlf_display_cols, 1):
            cell = ws.cell(row=tlf_header_row, column=current_col_idx)
            cell.value = make_visible_text_formula(col_name, date_ref, input_cell_ref)
            cell.font = header_font
            cell.alignment = align_center

        tlf_key_col = tlf_info["tlf_key_col_letter"]
        tlf_time_key_col = tlf_info["tlf_time_key_col_letter"]
        tlf_key_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${tlf_key_col}:${tlf_key_col}\")"
        tlf_time_key_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${tlf_time_key_col}:${tlf_time_key_col}\")"

        for r_offset in range(max_tlf_rows):
            k_value = r_offset + 1
            match_by_seq = f'MATCH("T|"&{input_cell_ref}&"|"&{k_value}, {tlf_key_range_str}, 0)'
            match_by_seq_and_time = f'MATCH("TT|"&{input_cell_ref}&"|"&{time_ref}&"|"&{k_value}, {tlf_time_key_range_str}, 0)'
            match_logic = f'IF({time_ref}="", {match_by_seq}, {match_by_seq_and_time})'
            helper_row = 2 + r_offset
            helper_match_formula = f'=IF(OR({input_cell_ref}="",{date_ref}=""), "", IFERROR({match_logic}*0+1, ""))'
            ws[f"{helper_match_col_letter}{helper_row}"] = helper_match_formula

        if TLF_TIME_HEADER in tlf_display_cols:
            time_col_idx = tlf_display_cols.index(TLF_TIME_HEADER) + 1
            time_col_letter = get_column_letter(time_col_idx)
            tlf_time_data_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${time_col_letter}:${time_col_letter}\")"

            for r_offset in range(max_tlf_rows):
                helper_row = 2 + r_offset
                k_value = r_offset + 1
                helper_formula = (
                    f'=IF(OR({input_cell_ref}="",{date_ref}=""), "", '
                    f'IFERROR(INDEX({tlf_time_data_range_str}, '
                    f'MATCH("T|"&{input_cell_ref}&"|"&{k_value}, {tlf_key_range_str}, 0)), ""))'
                )
                ws[f"{helper_time_col_letter}{helper_row}"] = helper_formula

            time_dv = DataValidation(
                type="list",
                formula1=f"=${helper_time_col_letter}$2:${helper_time_col_letter}${max_tlf_rows + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(time_dv)
            time_dv.add(time_cell)
    if max_tlf_rows > 0:
        ws[helper_visible_count_cell] = f'=COUNT(${helper_match_col_letter}$2:${helper_match_col_letter}${max_tlf_rows + 1})'
    else:
        ws[helper_visible_count_cell] = 0
    # ATMI starts right after the visible TLF block plus the configured gap.
    ws[helper_atmi_title_row_cell] = (
        f'=IF(OR({input_cell_ref}="",{date_ref}=""), 0, {tlf_data_start_row}+{gap_rows}+{helper_visible_count_ref})'
    )
    ws[helper_atmi_header_row_cell] = f'=IF({helper_atmi_title_row_ref}=0, 0, {helper_atmi_title_row_ref}+1)'

    atmi_title_min_row = tlf_data_start_row + gap_rows
    atmi_title_max_row = tlf_data_start_row + max_tlf_rows + gap_rows
    atmi_header_min_row = atmi_title_min_row + 1
    atmi_header_max_row = atmi_title_max_row + 1
    atmi_data_min_row = atmi_title_min_row + 2
    atmi_data_max_row = atmi_title_max_row + 1 + max_gl_rows

    search_body_end_row = max(tlf_data_start_row + max_tlf_rows - 1, atmi_data_max_row)

    tlf_key_range_str = None
    tlf_time_key_range_str = None
    if tlf_display_cols:
        tlf_key_col = tlf_info["tlf_key_col_letter"]
        tlf_time_key_col = tlf_info["tlf_time_key_col_letter"]
        tlf_key_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${tlf_key_col}:${tlf_key_col}\")"
        tlf_time_key_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${tlf_time_key_col}:${tlf_time_key_col}\")"

    gl_key_range_str = None
    if gl_display_cols:
        gl_key_col = gl_info["gl_key_col_letter"]
        gl_key_range_str = f"INDIRECT(\"'\"&{date_ref}&\"'!${gl_key_col}:${gl_key_col}\")"
        ws[helper_gl_visible_count_cell] = (
            f'=IF(OR({input_cell_ref}="",{date_ref}=""), 0, '
            f'COUNTIF({gl_key_range_str}, "G|"&{input_cell_ref}&"|*"))'
        )
    else:
        ws[helper_gl_visible_count_cell] = 0

    # Fill Search in one pass so overlapping sections do not overwrite each other.
    for current_row in range(tlf_data_start_row, search_body_end_row + 1):
        for out_col_idx in range(1, last_result_col_idx + 1):
            formula_parts = []

            if (
                current_row <= tlf_data_start_row + max_tlf_rows - 1
                and out_col_idx <= len(tlf_display_cols)
            ):
                tlf_col_letter = get_column_letter(out_col_idx)
                tlf_data_col_range = f"INDIRECT(\"'\"&{date_ref}&\"'!${tlf_col_letter}:${tlf_col_letter}\")"
                tlf_match_by_seq = (
                    f'MATCH("T|"&{input_cell_ref}&"|"&(ROW()-{tlf_data_start_row - 1}), {tlf_key_range_str}, 0)'
                )
                tlf_match_by_seq_and_time = (
                    f'MATCH("TT|"&{input_cell_ref}&"|"&{time_ref}&"|"&(ROW()-{tlf_data_start_row - 1}), '
                    f'{tlf_time_key_range_str}, 0)'
                )
                tlf_match_logic = f'IF({time_ref}="", {tlf_match_by_seq}, {tlf_match_by_seq_and_time})'
                tlf_data_formula = (
                    f'IF(AND(ROW()>={tlf_data_start_row}, ROW()<{tlf_data_start_row}+{helper_visible_count_ref}), '
                    f'IFERROR(INDEX({tlf_data_col_range}, {tlf_match_logic}), ""), "")'
                )
                if out_col_idx == 1:
                    tlf_formula = (
                        f'IF({helper_visible_count_ref}=0, '
                        f'IF(ROW()={tlf_data_start_row}, "ไม่พบข้อมูล", ""), {tlf_data_formula})'
                    )
                else:
                    tlf_formula = tlf_data_formula
                formula_parts.append(tlf_formula)

            if current_row >= atmi_title_min_row and out_col_idx <= len(gl_display_cols):
                gl_col_name = gl_display_cols[out_col_idx - 1]
                gl_col_letter = get_column_letter(out_col_idx)
                gl_data_col_range = f"INDIRECT(\"'\"&{date_ref}&\"'!${gl_col_letter}:${gl_col_letter}\")"
                header_label = SEARCH_HEADER_LABELS.get(gl_col_name, gl_col_name)
                escaped_col_name = str(header_label).replace('"', '""')
                gl_data_formula = (
                    f'IF(AND(ROW()>{helper_atmi_header_row_ref}, ROW()<={helper_atmi_header_row_ref}+{max_gl_rows}), '
                    f'IFERROR(INDEX({gl_data_col_range}, MATCH("G|"&{input_cell_ref}&"|"&(ROW()-{helper_atmi_header_row_ref}), {gl_key_range_str}, 0)), ""), "")'
                )

                if out_col_idx == 1:
                    gl_data_formula = (
                        f'IF({helper_gl_visible_count_ref}=0, '
                        f'IF(ROW()={helper_atmi_header_row_ref}+1, "ไม่พบข้อมูล", ""), {gl_data_formula})'
                    )
                    atmi_formula = (
                        f'IF(ROW()={helper_atmi_title_row_ref}, "ATMI", '
                        f'IF(ROW()={helper_atmi_header_row_ref}, "{escaped_col_name}", {gl_data_formula}))'
                    )
                else:
                    gl_data_formula = f'IF({helper_gl_visible_count_ref}=0, "", {gl_data_formula})'
                    atmi_formula = (
                        f'IF(ROW()={helper_atmi_header_row_ref}, "{escaped_col_name}", {gl_data_formula})'
                    )
                formula_parts.append(atmi_formula)

            cell = ws.cell(row=current_row, column=out_col_idx)
            if formula_parts:
                combined_formula = formula_parts[0]
                for extra_formula in formula_parts[1:]:
                    combined_formula = f'IF({combined_formula}<>"", {combined_formula}, {extra_formula})'
                cell.value = f'=IF(OR({input_cell_ref}="",{date_ref}=""), "", {combined_formula})'
            else:
                cell.value = None

            if current_row <= tlf_data_start_row + max_tlf_rows - 1 and out_col_idx <= len(tlf_display_cols):
                cell.alignment = align_center
                if out_col_idx == 9:
                    cell.number_format = "@"

            if current_row >= atmi_data_min_row and out_col_idx <= len(gl_display_cols):
                gl_col_name = gl_display_cols[out_col_idx - 1]
                if gl_col_name in ["DR", "CR"]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif gl_col_name == "Details":
                    cell.number_format = "@"
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
                    if gl_col_name == "Seq":
                        cell.number_format = "@"

    if tlf_display_cols and max_tlf_rows > 0:
        add_conditional_style(
            ws,
            f"A{tlf_header_row}:{tlf_last_col_letter}{tlf_header_row}",
            f'=$A{tlf_header_row}<>""',
            font=header_font,
            fill=header_fill,
            border=thin_border,
        )
        add_conditional_style(
            ws,
            f"A{tlf_data_start_row}:{tlf_last_col_letter}{tlf_data_start_row + max_tlf_rows - 1}",
            f'=${helper_match_col_letter}2=1',
            border=thin_border,
            alignment=align_center,
        )
        add_conditional_style(
            ws,
            f"A{tlf_data_start_row}:{tlf_last_col_letter}{tlf_data_start_row + max_tlf_rows - 1}",
            f'=$A{tlf_data_start_row}="ไม่พบข้อมูล"',
            fill=not_found_fill,
            border=thin_border,
            alignment=align_left,
        )
        add_conditional_style(
            ws,
            f"A{tlf_data_start_row}:A{tlf_data_start_row + max_tlf_rows - 1}",
            f'=$A{tlf_data_start_row}="ไม่พบข้อมูล"',
            font=not_found_font,
            alignment=align_left,
        )

    if gl_display_cols and max_gl_rows > 0:
        add_conditional_style(
            ws,
            f"A{atmi_title_min_row}:A{atmi_title_max_row}",
            f'=$A{atmi_title_min_row}="ATMI"',
            font=title_font,
            alignment=align_left,
        )
        add_conditional_style(
            ws,
            f"A{atmi_header_min_row}:{gl_last_col_letter}{atmi_header_max_row}",
            f'=ROW()={helper_atmi_header_row_ref}',
            font=header_font,
            fill=header_fill,
            border=thin_border,
        )
        add_conditional_style(
            ws,
            f"A{atmi_data_min_row}:{gl_last_col_letter}{atmi_data_max_row}",
            f'=AND(ROW()>{helper_atmi_header_row_ref}, ROW()<={helper_atmi_header_row_ref}+{max_gl_rows}, $A{atmi_data_min_row}<>"")',
            border=thin_border,
        )
        add_conditional_style(
            ws,
            f"A{atmi_data_min_row}:A{atmi_data_max_row}",
            f'=$A{atmi_data_min_row}="ไม่พบข้อมูล"',
            font=not_found_font,
            fill=not_found_fill,
            alignment=align_left,
        )

    merged_widths = merge_position_widths(
        *(info.get("tlf_search_widths", {}) for info in date_sheets_info),
        *(info.get("gl_search_widths", {}) for info in date_sheets_info),
    )
    fixed_width_positions = {}
    for idx, col_name in enumerate(tlf_display_cols, 1):
        if col_name in SEARCH_FIXED_WIDTHS:
            fixed_width_positions[idx] = max(fixed_width_positions.get(idx, 0), SEARCH_FIXED_WIDTHS[col_name])
    for idx, col_name in enumerate(gl_display_cols, 1):
        if col_name in SEARCH_FIXED_WIDTHS:
            fixed_width_positions[idx] = max(fixed_width_positions.get(idx, 0), SEARCH_FIXED_WIDTHS[col_name])

    for idx in range(1, last_result_col_idx + 1):
        width = fixed_width_positions.get(idx, merged_widths.get(idx, SEARCH_WIDTH_LIMITS[0]))
        if idx == 1:
            width = max(25, width)
        elif idx == 2:
            width = max(20, width)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.column_dimensions["AX"].hidden = True
    ws.column_dimensions[helper_match_col_letter].hidden = True
    ws.column_dimensions[helper_time_col_letter].hidden = True
    ws.column_dimensions["BA"].hidden = True


def load_tlf_books():
    books = []
    for filename in tlf_filenames:
        try:
            books.append((filename, pd.ExcelFile(filename)))
        except Exception as exc:
            print(f"Error อ่านไฟล์ TLF {filename}: {exc}")
            return None
    return books


def validate_inputs():
    if not os.path.exists(source_folder):
        print("Error: ไม่พบโฟลเดอร์หรือไฟล์ต้นฉบับ")
        return False

    missing_files = [filename for filename in tlf_filenames if not os.path.exists(filename)]
    if missing_files:
        print(f"Error: ไม่พบไฟล์ TLF ต่อไปนี้: {missing_files}")
        return False

    return True


def collect_chosen_dates(csv_files):
    return {file_path: pick_date_from_csv_col_c(file_path, nrows=20) for file_path in csv_files}


def process_file(writer, tlf_books, file_path, chosen_date, tlf_filter_val, tlf_reorder, pos_term_typ):
    filename = os.path.basename(file_path)
    desired_sheet_name = chosen_date if chosen_date else os.path.splitext(filename)[0]
    desired_sheet_name = sanitize_sheet_name(desired_sheet_name)
    print(f">> Processing: {filename} -> Sheet: {desired_sheet_name}")

    lookup_candidates = build_tlf_lookup_candidates(filename, chosen_date)
    tlf_sheet_to_use, tlf_book_to_use, tlf_book_filename = find_tlf_sheet(tlf_books, lookup_candidates)

    tlf_df = pd.DataFrame()
    effective_tlf_reserved_rows = tlf_reserved_rows
    max_k_tlf = 1

    if tlf_sheet_to_use:
        tlf_df, effective_tlf_reserved_rows, max_k_tlf = prepare_tlf_dataframe(
            tlf_book_to_use,
            tlf_sheet_to_use,
            tlf_filter_val,
            tlf_reorder,
            pos_term_typ,
            pos_AZ,
            pos_CU,
            pos_AF,
        )
        print(f"   ✓ TLF sheet used: {tlf_sheet_to_use} (file: {tlf_book_filename})")
    else:
        print(f"   ! ไม่พบชีตใน TLF (ลองแล้ว: {lookup_candidates}) -> ข้ามส่วน TLF")

    gl_df, effective_gl_reserved_rows, max_k_gl = prepare_gl_dataframe(file_path)

    target_sheet_name = make_unique_sheet_name(writer.book, desired_sheet_name)
    worksheet = writer.book.create_sheet(target_sheet_name)
    writer.sheets[target_sheet_name] = worksheet

    current_raw_row = 1
    tlf_key_col_letter = "A"
    tlf_time_key_col_letter = "A"
    gl_key_col_letter = "A"
    tlf_display_cols = []
    gl_display_cols = []

    if not tlf_df.empty:
        tlf_display_cols, tlf_key_col_letter, tlf_time_key_col_letter, current_raw_row = write_tlf_section(
            writer,
            target_sheet_name,
            worksheet,
            tlf_df,
            current_raw_row,
        )

    if not gl_df.empty:
        gl_display_cols, gl_key_col_letter = write_gl_section(
            writer,
            target_sheet_name,
            worksheet,
            gl_df,
            current_raw_row,
        )

    apply_column_widths(worksheet, tlf_df, tlf_display_cols, gl_df)

    tlf_search_widths = calculate_search_display_widths(tlf_df, tlf_display_cols)
    gl_search_widths = calculate_search_display_widths(gl_df, gl_display_cols)

    print(
        f"   ✓ เสร็จสิ้น: {filename} -> {target_sheet_name} | "
        f"chosen_date={chosen_date} | "
        f"TLF max|k={max_k_tlf} ui_rows={effective_tlf_reserved_rows} | "
        f"ATMI max|k={max_k_gl} ui_rows={effective_gl_reserved_rows}"
    )

    return {
        "sheet_name": target_sheet_name,
        "tlf_key_col_letter": tlf_key_col_letter,
        "tlf_time_key_col_letter": tlf_time_key_col_letter,
        "gl_key_col_letter": gl_key_col_letter,
        "tlf_display_cols": tlf_display_cols,
        "gl_display_cols": gl_display_cols,
        "tlf_search_widths": tlf_search_widths,
        "gl_search_widths": gl_search_widths,
        "effective_tlf_reserved_rows": effective_tlf_reserved_rows,
        "effective_gl_reserved_rows": effective_gl_reserved_rows,
    }


def process_combined_data():
    if not validate_inputs():
        return

    print("กำลังประมวลผล... (Auto Width & Reorder Column)")
    print(f"Source folder: {source_folder}")

    tlf_books = load_tlf_books()
    if tlf_books is None:
        return

    csv_files = list_csv_files(source_folder)
    if not csv_files:
        print("Error: ไม่พบไฟล์ .csv ในโฟลเดอร์")
        return

    chosen_dates = collect_chosen_dates(csv_files)

    print("-" * 30)
    print("ไฟล์ที่จะประมวลผล (หา date จาก Column C1-C20 แยกไฟล์):")
    for file_path in csv_files:
        filename = os.path.basename(file_path)
        print(f" - {filename} | chosen_date(YYMMDD)={chosen_dates[file_path]}")
    print("-" * 30)

    folder_upper = os.path.basename(os.path.normpath(source_folder)).upper()
    tlf_filter_val = TERM_TYPE_FILTERS.get(folder_upper)

    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        date_sheets_info = []
        for file_path in csv_files:
            try:
                date_sheets_info.append(
                    process_file(
                        writer,
                        tlf_books,
                        file_path,
                        chosen_dates[file_path],
                        tlf_filter_val,
                        tlf_reorder,
                        pos_term_typ,
                    )
                )
            except Exception as exc:
                print(f"X Error ไฟล์ {os.path.basename(file_path)}: {exc}")
                import traceback

                traceback.print_exc()

        create_search_sheet(writer, date_sheets_info)

        if "Search" in writer.book.sheetnames:
            search_sheet = writer.book["Search"]
            writer.book._sheets.remove(search_sheet)
            writer.book._sheets.insert(0, search_sheet)

        if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
            del writer.book["Sheet"]

    print("-" * 30)
    print(f"บันทึกไฟล์เรียบร้อยที่: {output_filename}")


def pick_date_from_csv_col_c_bytes(file_bytes: bytes, nrows=20):
    """Pick the most frequent date from Column C and return it as YYMMDD."""
    df_col_c = None
    for encoding in CSV_ENCODINGS:
        try:
            df_col_c = pd.read_csv(
                io.BytesIO(file_bytes),
                header=None,
                usecols=[2],
                nrows=nrows,
                encoding=encoding,
                dtype=str,
                engine="python",
            )
            break
        except Exception:
            df_col_c = None

    if df_col_c is None or df_col_c.empty:
        return None

    date_values = [to_yymmdd(value) for value in df_col_c.iloc[:, 0].tolist()]
    date_values = [value for value in date_values if value]
    if not date_values:
        return None

    value_counts = pd.Series(date_values).value_counts()
    max_count = value_counts.max()
    candidates = [date_value for date_value, count in value_counts.items() if count == max_count]
    return max(candidates, key=lambda item: int(item))


def read_gl_csv_bytes(file_bytes: bytes, filename_for_error: str):
    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(
                io.BytesIO(file_bytes),
                header=None,
                usecols=gl_indices,
                encoding=encoding,
                dtype=str,
                engine="python",
            )
        except Exception:
            continue
    raise ValueError(f"ไม่สามารถอ่านไฟล์ด้วย encoding ที่รองรับได้: {filename_for_error}")


def load_tlf_books(tlf_folder_path: str = DEFAULT_TLF_FOLDER):
    if not os.path.isdir(tlf_folder_path):
        raise ValueError(f"ไม่พบ folder TLF: {tlf_folder_path}")

    tlf_files = sorted(
        [
            filename
            for filename in os.listdir(tlf_folder_path)
            if filename.lower().endswith(".xlsx") and not filename.startswith("~$")
        ],
        key=lambda name: name.lower(),
    )
    if not tlf_files:
        raise ValueError(f"ไม่พบไฟล์ TLF (.xlsx) ใน folder: {tlf_folder_path}")

    books = []
    for filename in tlf_files:
        full_path = os.path.join(tlf_folder_path, filename)
        try:
            books.append((filename, pd.ExcelFile(full_path)))
        except Exception as exc:
            raise RuntimeError(f"Error อ่านไฟล์ TLF {filename}: {exc}") from exc
    return books


def prepare_gl_dataframe_from_bytes(file_bytes: bytes, filename_for_error: str):
    gl_df = read_gl_csv_bytes(file_bytes, filename_for_error)

    if len(gl_df.columns) == len(gl_source_headers):
        gl_df.columns = gl_source_headers

    gl_df["Details"] = gl_df["AZ_RAW"]
    gl_df["Seq"] = gl_df["AZ_RAW"].apply(extract_seq_num).astype(str).str.strip()

    gl_df["RC"] = gl_df["RC"].astype(str).str.strip()
    gl_df["CH"] = gl_df["CH"].astype(str).str.strip()
    gl_df["DR"] = pd.to_numeric(gl_df["DR"], errors="coerce").fillna(0)
    gl_df["CR"] = pd.to_numeric(gl_df["CR"], errors="coerce").fillna(0)

    gl_df = gl_df[gl_base_headers]
    gl_df = gl_df.sort_values(
        by=["CH", "RC", "OC", "Product Code"],
        ascending=[True, True, True, True],
    )

    if not gl_df.empty:
        search_col_gl = gl_df["Seq"].astype(str)
        gl_df["_SearchKey"] = "G|" + search_col_gl + "|" + (gl_df.groupby(search_col_gl).cumcount() + 1).astype(str)

    max_k_gl = 1
    if not gl_df.empty and "_SearchKey" in gl_df.columns:
        max_k_gl = max_k_from_searchkey(gl_df["_SearchKey"])

    effective_gl_reserved_rows = max(gl_reserved_rows, max_k_gl)
    return gl_df, effective_gl_reserved_rows, max_k_gl


def process_file_from_zip(writer, tlf_books, filename, file_bytes, chosen_date, tlf_filter_val, log):
    desired_sheet_name = chosen_date if chosen_date else os.path.splitext(filename)[0]
    desired_sheet_name = sanitize_sheet_name(desired_sheet_name)
    log(f">> Processing: {filename} -> Sheet: {desired_sheet_name}")

    lookup_candidates = build_tlf_lookup_candidates(filename, chosen_date)
    tlf_sheet_to_use, tlf_book_to_use, tlf_book_filename = find_tlf_sheet(tlf_books, lookup_candidates)

    tlf_df = pd.DataFrame()
    effective_tlf_reserved_rows = tlf_reserved_rows
    max_k_tlf = 1

    if tlf_sheet_to_use:
        tlf_df, effective_tlf_reserved_rows, max_k_tlf = prepare_tlf_dataframe(
            tlf_book_to_use,
            tlf_sheet_to_use,
            tlf_filter_val,
            tlf_reorder,
            pos_term_typ,
            pos_AZ,
            pos_CU,
            pos_AF,
        )
        log(f"   ✓ TLF sheet used: {tlf_sheet_to_use} (file: {tlf_book_filename})")
    else:
        log(f"   ! ไม่พบชีตใน TLF (ลองแล้ว: {lookup_candidates}) -> ข้ามส่วน TLF")

    gl_df, effective_gl_reserved_rows, max_k_gl = prepare_gl_dataframe_from_bytes(file_bytes, filename)

    target_sheet_name = make_unique_sheet_name(writer.book, desired_sheet_name)
    worksheet = writer.book.create_sheet(target_sheet_name)
    writer.sheets[target_sheet_name] = worksheet

    current_raw_row = 1
    tlf_key_col_letter = "A"
    tlf_time_key_col_letter = "A"
    gl_key_col_letter = "A"
    tlf_display_cols = []
    gl_display_cols = []

    if not tlf_df.empty:
        tlf_display_cols, tlf_key_col_letter, tlf_time_key_col_letter, current_raw_row = write_tlf_section(
            writer,
            target_sheet_name,
            worksheet,
            tlf_df,
            current_raw_row,
        )

    if not gl_df.empty:
        gl_display_cols, gl_key_col_letter = write_gl_section(
            writer,
            target_sheet_name,
            worksheet,
            gl_df,
            current_raw_row,
        )

    apply_column_widths(worksheet, tlf_df, tlf_display_cols, gl_df)

    tlf_search_widths = calculate_search_display_widths(tlf_df, tlf_display_cols)
    gl_search_widths = calculate_search_display_widths(gl_df, gl_display_cols)

    log(
        f"   ✓ เสร็จสิ้น: {filename} -> {target_sheet_name} | "
        f"chosen_date={chosen_date} | "
        f"TLF max|k={max_k_tlf} ui_rows={effective_tlf_reserved_rows} | "
        f"ATMI max|k={max_k_gl} ui_rows={effective_gl_reserved_rows}"
    )

    return {
        "sheet_name": target_sheet_name,
        "tlf_key_col_letter": tlf_key_col_letter,
        "tlf_time_key_col_letter": tlf_time_key_col_letter,
        "gl_key_col_letter": gl_key_col_letter,
        "tlf_display_cols": tlf_display_cols,
        "gl_display_cols": gl_display_cols,
        "tlf_search_widths": tlf_search_widths,
        "gl_search_widths": gl_search_widths,
        "effective_tlf_reserved_rows": effective_tlf_reserved_rows,
        "effective_gl_reserved_rows": effective_gl_reserved_rows,
    }


def process_combined_data_from_zip(
    zip_bytes: bytes,
    source_folder_name: str,
    tlf_folder_path: str = DEFAULT_TLF_FOLDER,
):
    output_filename = f"GL_{source_folder_name}.xlsx"
    log_lines = []

    def log(message: str):
        log_lines.append(message)

    tlf_books = load_tlf_books(tlf_folder_path)

    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        names = [name for name in zf.namelist() if not name.endswith("/")]
        flat_names = [name for name in names if "/" not in name and "\\" not in name]
        csv_files = sorted([name for name in flat_names if name.lower().endswith(".csv")], key=lambda name: name.lower())

        if not csv_files:
            raise ValueError("ไม่พบไฟล์ .csv ใน ZIP (ต้องอยู่ระดับเดียวกัน ไม่อยู่ใน folder)")

        csv_bytes_map = {filename: zf.read(filename) for filename in csv_files}
        chosen_dates = {filename: pick_date_from_csv_col_c_bytes(file_bytes, nrows=20) for filename, file_bytes in csv_bytes_map.items()}

    selected_source_type, tlf_filter_val = get_term_type_from_source_name(source_folder_name)

    log("กำลังประมวลผล... (ZIP workflow + Excel format แบบ ex.py)")
    log(f"Source folder: {source_folder_name}")
    log(f"TLF folder: {tlf_folder_path}")
    log(f"TLF files: {[filename for filename, _ in tlf_books]}")
    log(f"CSV files: {csv_files}")
    if selected_source_type and tlf_filter_val:
        log(f"Selected source type: {selected_source_type} -> TERM TYPE {tlf_filter_val}")
    else:
        log("Selected source type: none -> ไม่กรอง TERM TYPE")
    log("-" * 30)
    log("ไฟล์ที่จะประมวลผล (หา date จาก Column C1-C20 แยกไฟล์):")
    for filename in csv_files:
        log(f" - {filename} | chosen_date(YYMMDD)={chosen_dates[filename]}")
    log("-" * 30)

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        date_sheets_info = []
        for filename in csv_files:
            try:
                date_sheets_info.append(
                    process_file_from_zip(
                        writer,
                        tlf_books,
                        filename,
                        csv_bytes_map[filename],
                        chosen_dates[filename],
                        tlf_filter_val,
                        log,
                    )
                )
            except Exception as exc:
                log(f"X Error ไฟล์ {filename}: {exc}")
                log(traceback.format_exc())

        create_search_sheet(writer, date_sheets_info)

        if "Search" in writer.book.sheetnames:
            search_sheet = writer.book["Search"]
            writer.book._sheets.remove(search_sheet)
            writer.book._sheets.insert(0, search_sheet)

        if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
            del writer.book["Sheet"]

    out.seek(0)
    log("-" * 30)
    log(f"บันทึกไฟล์เรียบร้อยที่: {output_filename}")

    return out, output_filename, "\n".join(log_lines)


def render():
    st.write("อัปโหลด ZIP ที่มีเฉพาะไฟล์ **.csv** และระบบจะดึงไฟล์ TLF จาก folder **Data GL** ให้อัตโนมัติ")

    source_name = st.selectbox(
        "Source name (ระบุเพื่อแบ่งข้อมูล ATM/CDM และใช้ตั้งชื่อไฟล์ output: GL_<source>.xlsx)",
        options=["ATM", "CDM"],
        index=1 if DEFAULT_SOURCE_FOLDER_NAME == "CDM" else 0,
    )
    uploaded = st.file_uploader("Upload ZIP (เฉพาะ .csv)", type=["zip"])

    if uploaded is not None:
        try:
            out_bytes, out_name, logs = process_combined_data_from_zip(
                uploaded.getvalue(),
                source_name,
                DEFAULT_TLF_FOLDER,
            )

            st.success("✅ ประมวลผลเสร็จสิ้น")
            st.download_button(
                label=f"⬇️ ดาวน์โหลดไฟล์ {out_name}",
                data=out_bytes.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            with st.expander("ดู Log"):
                st.code(logs)

        except Exception as exc:
            st.error(f"❌ Error: {exc}")


gl_indices = [excel_col_to_index(col) for col in gl_columns_letters]
tlf_indices = [excel_col_to_index(col) for col in tlf_columns_letters]
tlf_reorder = [sorted(tlf_indices).index(idx) for idx in tlf_indices]
pos_AZ = get_col_pos_in_tlf("AZ")
pos_CU = get_col_pos_in_tlf("CU")
pos_AF = get_col_pos_in_tlf("AF")
pos_term_typ = get_col_pos_in_tlf("M")

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
header_font = Font(bold=True)
title_font = Font(bold=True, size=14, color="000000")
not_found_font = Font(bold=True, size=14, color="C00000")
search_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
not_found_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")


if __name__ == "__main__":
    configure_stdout()
    print("Use this module through the Streamlit app.")
